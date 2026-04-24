"""
ashrah_backfill.py — Lumia, Ashrah Painting Operations Agent
=============================================================
7 modules:
  1. Email Scanner          — fetch & parse inbound emails via Zoho IMAP
  2. Job Analysis           — extract structured job data from email threads
  3. Workforce Tracking     — maintain painter/crew availability and history
  4. QC Inspector Comparison — diff inspector reports and flag discrepancies
  5. Site Assignment        — match crew to open jobs based on skills/availability
  6. Client Daily Report    — compose and send polished site reports via Zoho SMTP
  7. Employee Daily Log     — record each employee's site, work done, and self-score to Excel

Requires:
  pip install anthropic openpyxl
  Environment variables (see CONFIG section below)
"""

from __future__ import annotations

import imaplib
import smtplib
import email
import json
import os
import re
import textwrap
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import decode_header
from typing import Any

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG — set these as environment variables or fill in defaults
# ---------------------------------------------------------------------------
ZOHO_IMAP_HOST    = os.getenv("ZOHO_IMAP_HOST",    "imap.zoho.com")
ZOHO_IMAP_PORT    = int(os.getenv("ZOHO_IMAP_PORT", "993"))
ZOHO_SMTP_HOST    = os.getenv("ZOHO_SMTP_HOST",    "smtp.zoho.com")
ZOHO_SMTP_PORT    = int(os.getenv("ZOHO_SMTP_PORT", "465"))
ZOHO_EMAIL        = os.getenv("ZOHO_EMAIL",         "ops@ashrahpainting.com")
ZOHO_PASSWORD     = os.getenv("ZOHO_PASSWORD",      "")   # Zoho app password
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY",  "")

MODEL = "claude-opus-4-6"

# ---------------------------------------------------------------------------
# DATA MODELS
# ---------------------------------------------------------------------------

@dataclass
class EmailMessage:
    uid: str
    subject: str
    sender: str
    date_str: str
    body: str
    thread_id: str = ""


@dataclass
class Job:
    job_id: str
    client_name: str
    site_address: str
    description: str
    start_date: str = ""
    end_date: str = ""
    status: str = "open"           # open | in_progress | complete | on_hold
    assigned_crew: list[str] = field(default_factory=list)
    notes: str = ""
    raw_source: str = ""


@dataclass
class Worker:
    worker_id: str
    name: str
    skills: list[str] = field(default_factory=list)   # interior, exterior, spray, etc.
    available: bool = True
    current_site: str = ""
    phone: str = ""
    email: str = ""
    certifications: list[str] = field(default_factory=list)


@dataclass
class QCReport:
    report_id: str
    job_id: str
    inspector_name: str
    inspection_date: str
    items: dict[str, str] = field(default_factory=dict)   # area -> pass/fail/note
    overall_rating: str = ""
    raw_text: str = ""


@dataclass
class SiteAssignment:
    job_id: str
    site_address: str
    assigned_workers: list[str]
    assignment_date: str
    notes: str = ""


@dataclass
class EmployeeDailyEntry:
    entry_date: str           # YYYY-MM-DD
    worker_id: str
    worker_name: str
    site_address: str
    job_id: str
    work_description: str     # daily summary
    self_score: int           # average of category scores (1-10)
    notes: str = ""
    # Category scores (1-10 each)
    tape_covering: int = 0
    drop_sheets: int = 0
    patching_process: int = 0
    paint_execution: int = 0
    site_control: int = 0
    washing_tool_care: int = 0
    # Optional employee-defined scores
    custom_scores: str = ""   # e.g. "Primer: 9 | Window Trim: 8"
    # Tomorrow's plan
    tomorrows_plan: str = ""


@dataclass
class DailyReport:
    report_date: str
    job_id: str
    site_address: str
    client_name: str
    client_email: str
    crew_present: list[str]
    work_completed: str
    work_planned: str
    issues: str
    photos_pending: int = 0
    overall_status: str = "On Schedule"


# ---------------------------------------------------------------------------
# MODULE 1 — EMAIL SCANNER
# ---------------------------------------------------------------------------

class EmailScanner:
    """Connect to Zoho IMAP, fetch unread messages, return EmailMessage list."""

    def __init__(self, host: str, port: int, user: str, password: str):
        self.host = host
        self.port = port
        self.user = user
        self.password = password

    def _decode_header_value(self, value: str) -> str:
        parts = decode_header(value)
        decoded = []
        for part, charset in parts:
            if isinstance(part, bytes):
                decoded.append(part.decode(charset or "utf-8", errors="replace"))
            else:
                decoded.append(part)
        return " ".join(decoded)

    def _extract_body(self, msg: email.message.Message) -> str:
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                ct = part.get_content_type()
                cd = str(part.get("Content-Disposition", ""))
                if ct == "text/plain" and "attachment" not in cd:
                    payload = part.get_payload(decode=True)
                    if payload:
                        body += payload.decode(
                            part.get_content_charset() or "utf-8", errors="replace"
                        )
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode(
                    msg.get_content_charset() or "utf-8", errors="replace"
                )
        return body.strip()

    def fetch_unread(self, mailbox: str = "INBOX", max_count: int = 50) -> list[EmailMessage]:
        messages: list[EmailMessage] = []
        try:
            conn = imaplib.IMAP4_SSL(self.host, self.port)
            conn.login(self.user, self.password)
            conn.select(mailbox)

            _, data = conn.search(None, "UNSEEN")
            uids = data[0].split()
            if not uids:
                print("[EmailScanner] No unread messages.")
                conn.logout()
                return messages

            for uid in uids[-max_count:]:
                _, msg_data = conn.fetch(uid, "(RFC822)")
                if not msg_data or not msg_data[0]:
                    continue
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)

                subject  = self._decode_header_value(msg.get("Subject", ""))
                sender   = self._decode_header_value(msg.get("From", ""))
                date_str = msg.get("Date", "")
                body     = self._extract_body(msg)
                thread   = msg.get("Message-ID", uid.decode())

                messages.append(EmailMessage(
                    uid=uid.decode(),
                    subject=subject,
                    sender=sender,
                    date_str=date_str,
                    body=body,
                    thread_id=thread,
                ))

            conn.logout()
        except imaplib.IMAP4.error as exc:
            print(f"[EmailScanner] IMAP error: {exc}")
        return messages

    def mark_as_read(self, uid: str, mailbox: str = "INBOX") -> None:
        try:
            conn = imaplib.IMAP4_SSL(self.host, self.port)
            conn.login(self.user, self.password)
            conn.select(mailbox)
            conn.store(uid, "+FLAGS", "\\Seen")
            conn.logout()
        except imaplib.IMAP4.error as exc:
            print(f"[EmailScanner] Could not mark UID {uid} as read: {exc}")


# ---------------------------------------------------------------------------
# MODULE 2 — JOB ANALYSIS
# ---------------------------------------------------------------------------

class JobAnalyzer:
    """
    Use Claude to extract structured job data from raw email content.
    Returns a list of Job objects.
    """

    SYSTEM_PROMPT = textwrap.dedent("""
        You are an operations analyst for Ashrah Painting, a commercial and residential
        painting contractor. Your task is to parse incoming emails and extract painting
        job details into a strict JSON format.

        Always respond with a JSON array. Each element represents one distinct job found
        in the email. If no job is found, return an empty array [].

        Job object schema:
        {
          "job_id":       "<generate a short slug like JOB-2026-001 if none provided>",
          "client_name":  "<full client name>",
          "site_address": "<full site address>",
          "description":  "<scope of painting work>",
          "start_date":   "<YYYY-MM-DD or empty string>",
          "end_date":     "<YYYY-MM-DD or empty string>",
          "status":       "<open|in_progress|complete|on_hold>",
          "notes":        "<any special instructions, access codes, surface prep notes>"
        }

        Be concise. Extract only what is clearly stated.
    """).strip()

    def __init__(self, client: anthropic.Anthropic):
        self.client = client

    def _analyze_one(self, msg: EmailMessage) -> list[Job]:
        content = f"FROM: {msg.sender}\nSUBJECT: {msg.subject}\nDATE: {msg.date_str}\n\n{msg.body[:3000]}"
        try:
            with self.client.messages.stream(
                model=MODEL,
                max_tokens=1024,
                system=self.SYSTEM_PROMPT,
                messages=[{"role": "user", "content": content}],
            ) as stream:
                response = stream.get_final_message()
        except Exception as exc:
            print(f"[JobAnalyzer] API error on UID {msg.uid}: {exc}")
            return []

        raw = next((b.text for b in response.content if b.type == "text"), "[]")
        raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()
        try:
            parsed = json.loads(raw)
            return [
                Job(
                    job_id=item.get("job_id", f"JOB-{date.today().year}-UNK"),
                    client_name=item.get("client_name", "Unknown Client"),
                    site_address=item.get("site_address", ""),
                    description=item.get("description", ""),
                    start_date=item.get("start_date", ""),
                    end_date=item.get("end_date", ""),
                    status=item.get("status", "open"),
                    notes=item.get("notes", ""),
                    raw_source=content[:500],
                )
                for item in parsed
            ]
        except json.JSONDecodeError as exc:
            print(f"[JobAnalyzer] JSON parse error: {exc}\nRaw: {raw[:300]}")
            return []

    def analyze(self, messages: list[EmailMessage]) -> list[Job]:
        jobs: list[Job] = []
        for msg in messages:
            jobs.extend(self._analyze_one(msg))
        return jobs


# ---------------------------------------------------------------------------
# MODULE 2b — EMPLOYEE EMAIL PARSER
# ---------------------------------------------------------------------------

class EmployeeEmailParser:
    """
    Reads inbound emails from employees and extracts:
      - Their name
      - Site address they worked at
      - What they did
      - Self-score (1-10)
    """

    SYSTEM_PROMPT = textwrap.dedent("""
        You are reading a daily check-in email sent by a painter at Ashrah Painting.
        Extract the following fields and return ONLY valid JSON — no markdown fences.

        {
          "worker_name":        "<full name of the sender, or 'Unknown' if not found>",
          "site_address":       "<full address of the site they worked at, or 'Unknown' if not found>",
          "work_description":   "<what they did today — quote or paraphrase from the email>",
          "self_score":         <integer 1-10, or 0 if not mentioned>,
          "notes":              "<anything extra they mentioned>"
        }

        If the email is clearly NOT a daily check-in (e.g. spam, out-of-office, client inquiry),
        return exactly: {}
    """).strip()

    def __init__(self, client: anthropic.Anthropic):
        self.client = client

    def parse_entries(self, messages: list[EmailMessage]) -> list[EmployeeDailyEntry]:
        entries: list[EmployeeDailyEntry] = []
        for msg in messages:
            content = (
                f"FROM: {msg.sender}\n"
                f"SUBJECT: {msg.subject}\n"
                f"DATE: {msg.date_str}\n\n"
                f"{msg.body[:2000]}"
            )
            try:
                with self.client.messages.stream(
                    model=MODEL,
                    max_tokens=512,
                    system=self.SYSTEM_PROMPT,
                    messages=[{"role": "user", "content": content}],
                ) as stream:
                    response = stream.get_final_message()
            except Exception as exc:
                print(f"[EmployeeEmailParser] API error: {exc}")
                continue

            raw = next((b.text for b in response.content if b.type == "text"), "{}")
            raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()

            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                continue

            if not data:
                continue

            # Try to parse the date from the email header
            try:
                entry_date = datetime.strptime(
                    msg.date_str[:16].strip(), "%a, %d %b %Y"
                ).strftime("%Y-%m-%d")
            except ValueError:
                entry_date = date.today().isoformat()

            score = int(data.get("self_score", 0))
            if not (1 <= score <= 10):
                score = 0

            name = data.get("worker_name", "Unknown")
            site = data.get("site_address", "Unknown")

            if name == "Unknown" and site == "Unknown":
                continue   # not a real employee check-in

            entries.append(EmployeeDailyEntry(
                entry_date=entry_date,
                worker_id="",
                worker_name=name,
                site_address=site,
                job_id="",
                work_description=data.get("work_description", ""),
                self_score=score,
                notes=data.get("notes", ""),
            ))
            print(f"[EmployeeEmailParser] Parsed entry: {name} @ {site} — score {score}")

        return entries


# ---------------------------------------------------------------------------
# MODULE 3 — WORKFORCE TRACKING
# ---------------------------------------------------------------------------

class WorkforceTracker:
    """
    In-memory workforce registry. In production, back this with a database
    or a JSON file persisted to disk.
    """

    def __init__(self):
        self._workers: dict[str, Worker] = {}

    # ---- CRUD ----------------------------------------------------------------

    def add_worker(self, worker: Worker) -> None:
        self._workers[worker.worker_id] = worker
        print(f"[WorkforceTracker] Added: {worker.name} ({worker.worker_id})")

    def update_availability(self, worker_id: str, available: bool, site: str = "") -> None:
        if worker_id not in self._workers:
            print(f"[WorkforceTracker] Unknown worker ID: {worker_id}")
            return
        w = self._workers[worker_id]
        w.available = available
        w.current_site = site if not available else ""
        print(f"[WorkforceTracker] {w.name}: available={available}, site='{site}'")

    def get_available(self, skill: str = "") -> list[Worker]:
        pool = [w for w in self._workers.values() if w.available]
        if skill:
            pool = [w for w in pool if skill.lower() in [s.lower() for s in w.skills]]
        return pool

    def get_all(self) -> list[Worker]:
        return list(self._workers.values())

    def summary(self) -> str:
        lines = ["WORKFORCE SUMMARY", "-" * 50]
        for w in self._workers.values():
            status = "AVAILABLE" if w.available else f"ON SITE: {w.current_site}"
            skills = ", ".join(w.skills) or "—"
            lines.append(f"  {w.name:20s}  [{status}]  Skills: {skills}")
        lines.append("-" * 50)
        total     = len(self._workers)
        available = sum(1 for w in self._workers.values() if w.available)
        lines.append(f"Total: {total}  Available: {available}  On site: {total - available}")
        return "\n".join(lines)

    # ---- AI-assisted analysis -----------------------------------------------

    def analyze_with_claude(self, client: anthropic.Anthropic, question: str) -> str:
        """Ask Claude a workforce question given current state."""
        workforce_data = json.dumps(
            [asdict(w) for w in self._workers.values()], indent=2
        )
        system = textwrap.dedent("""
            You are the workforce coordinator for Ashrah Painting.
            Answer workforce questions based on the provided JSON crew data.
            Be concise and actionable.
        """).strip()
        user_content = f"Crew data:\n{workforce_data}\n\nQuestion: {question}"

        with client.messages.stream(
            model=MODEL,
            max_tokens=1024,
            system=system,
            messages=[{"role": "user", "content": user_content}],
        ) as stream:
            response = stream.get_final_message()

        return next((b.text for b in response.content if b.type == "text"), "")


# ---------------------------------------------------------------------------
# MODULE 4 — QC INSPECTOR COMPARISON
# ---------------------------------------------------------------------------

class QCInspectorComparison:
    """
    Compare multiple QC inspection reports for the same job.
    Claude flags discrepancies, scores alignment, and recommends actions.
    """

    SYSTEM_PROMPT = textwrap.dedent("""
        You are a quality control manager for Ashrah Painting.
        Given two or more QC inspection reports for the same job site,
        compare them and produce a structured JSON response:

        {
          "job_id": "<job id>",
          "inspection_date": "<date>",
          "agreement_score": <0-100>,
          "discrepancies": [
            {
              "area": "<area name>",
              "inspector_a": "<rating/note from inspector A>",
              "inspector_b": "<rating/note from inspector B>",
              "severity": "<low|medium|high>",
              "recommended_action": "<what to do>"
            }
          ],
          "consensus_issues": ["<issue that both inspectors flagged>"],
          "summary": "<2-3 sentence executive summary>"
        }

        Respond only with valid JSON. Do not wrap in markdown fences.
    """).strip()

    def __init__(self, client: anthropic.Anthropic):
        self.client = client

    def compare(self, reports: list[QCReport]) -> dict[str, Any]:
        if len(reports) < 2:
            print("[QCComparison] Need at least 2 reports to compare.")
            return {}

        report_texts = []
        for i, r in enumerate(reports, 1):
            block = (
                f"=== Inspector {i}: {r.inspector_name} | {r.inspection_date} ===\n"
                f"Overall rating: {r.overall_rating}\n"
            )
            for area, note in r.items.items():
                block += f"  {area}: {note}\n"
            if r.raw_text:
                block += f"\nFull notes:\n{r.raw_text}\n"
            report_texts.append(block)

        payload = (
            f"Job ID: {reports[0].job_id}\n\n"
            + "\n\n".join(report_texts)
        )

        with self.client.messages.stream(
            model=MODEL,
            max_tokens=2048,
            thinking={"type": "adaptive"},
            system=self.SYSTEM_PROMPT,
            messages=[{"role": "user", "content": payload}],
        ) as stream:
            response = stream.get_final_message()

        raw = next((b.text for b in response.content if b.type == "text"), "{}")
        raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()

        try:
            return json.loads(raw)
        except json.JSONDecodeError as exc:
            print(f"[QCComparison] JSON parse error: {exc}")
            return {"raw": raw}

    def format_comparison(self, result: dict[str, Any]) -> str:
        if not result:
            return "No comparison result."
        lines = [
            f"QC COMPARISON — Job {result.get('job_id', '?')}",
            f"Agreement Score : {result.get('agreement_score', '?')}/100",
            f"Summary         : {result.get('summary', '')}",
            "",
        ]
        discrepancies = result.get("discrepancies", [])
        if discrepancies:
            lines.append(f"DISCREPANCIES ({len(discrepancies)}):")
            for d in discrepancies:
                lines.append(
                    f"  [{d.get('severity','?').upper():6s}] {d.get('area','')}: "
                    f"A='{d.get('inspector_a','')}' vs B='{d.get('inspector_b','')}'"
                    f"\n           Action: {d.get('recommended_action','')}"
                )
        consensus = result.get("consensus_issues", [])
        if consensus:
            lines.append("\nCONSENSUS ISSUES (both inspectors flagged):")
            for c in consensus:
                lines.append(f"  • {c}")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# MODULE 5 — SITE ASSIGNMENT
# ---------------------------------------------------------------------------

class SiteAssigner:
    """
    Use Claude to recommend optimal crew assignments for open jobs,
    then persist the assignments and update workforce availability.
    """

    SYSTEM_PROMPT = textwrap.dedent("""
        You are the site operations coordinator for Ashrah Painting.
        Given a list of open painting jobs and available workers, assign
        the best crew to each job.

        Rules:
        - Match worker skills to job requirements (exterior, interior, spray, drywall, etc.).
        - Do not assign a worker to more than one job at a time.
        - Prefer workers whose certifications match the job requirements
          (e.g., lead-safe certification for pre-1978 buildings).
        - Aim for crews of 2-4 per residential job, 4-8 per commercial job.
        - Leave at least one worker unassigned as a flex/backup if possible.

        Respond with a JSON array of assignments:
        [
          {
            "job_id": "<job id>",
            "assigned_workers": ["<worker_id>", ...],
            "assignment_notes": "<brief rationale>"
          }
        ]

        Respond only with valid JSON. Do not wrap in markdown fences.
    """).strip()

    def __init__(self, client: anthropic.Anthropic, tracker: WorkforceTracker):
        self.client = client
        self.tracker = tracker
        self._assignments: list[SiteAssignment] = []

    def assign(self, jobs: list[Job]) -> list[SiteAssignment]:
        open_jobs = [j for j in jobs if j.status == "open"]
        if not open_jobs:
            print("[SiteAssigner] No open jobs to assign.")
            return []

        available = self.tracker.get_available()
        if not available:
            print("[SiteAssigner] No available workers.")
            return []

        jobs_json    = json.dumps([asdict(j) for j in open_jobs], indent=2)
        workers_json = json.dumps([asdict(w) for w in available], indent=2)
        payload      = f"Open jobs:\n{jobs_json}\n\nAvailable workers:\n{workers_json}"

        with self.client.messages.stream(
            model=MODEL,
            max_tokens=2048,
            thinking={"type": "adaptive"},
            system=self.SYSTEM_PROMPT,
            messages=[{"role": "user", "content": payload}],
        ) as stream:
            response = stream.get_final_message()

        raw = next((b.text for b in response.content if b.type == "text"), "[]")
        raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()

        assignments: list[SiteAssignment] = []
        try:
            parsed = json.loads(raw)
            for item in parsed:
                jid     = item.get("job_id", "")
                workers = item.get("assigned_workers", [])
                notes   = item.get("assignment_notes", "")

                job  = next((j for j in open_jobs if j.job_id == jid), None)
                site = job.site_address if job else ""

                a = SiteAssignment(
                    job_id=jid,
                    site_address=site,
                    assigned_workers=workers,
                    assignment_date=date.today().isoformat(),
                    notes=notes,
                )
                assignments.append(a)

                # Update job state
                if job:
                    job.assigned_crew = workers
                    job.status = "in_progress"

                # Mark workers as unavailable
                for wid in workers:
                    self.tracker.update_availability(wid, available=False, site=site)

        except json.JSONDecodeError as exc:
            print(f"[SiteAssigner] JSON parse error: {exc}")

        self._assignments.extend(assignments)
        return assignments

    def get_assignments(self) -> list[SiteAssignment]:
        return self._assignments

    def summary(self) -> str:
        if not self._assignments:
            return "No site assignments recorded."
        lines = ["SITE ASSIGNMENTS", "=" * 55]
        for a in self._assignments:
            lines.append(f"  Job     : {a.job_id}")
            lines.append(f"  Site    : {a.site_address}")
            lines.append(f"  Workers : {', '.join(a.assigned_workers)}")
            lines.append(f"  Date    : {a.assignment_date}")
            if a.notes:
                lines.append(f"  Notes   : {a.notes}")
            lines.append("")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# MODULE 6 — CLIENT-FACING DAILY SITE REPORT
# ---------------------------------------------------------------------------

class DailyReportSender:
    """
    Generate a polished HTML + plain-text daily site report with Claude,
    then send it via Zoho SMTP.
    """

    SYSTEM_PROMPT = textwrap.dedent("""
        You are writing a daily site report for Ashrah Painting. This email goes directly to the client.

        ## WRITING RULES — READ CAREFULLY

        VOICE & ATTRIBUTION:
        - Always write as "the team" — never name individual crew members or say who did what.
          Wrong: "Weas taped the baseboards and Ismael patched the walls."
          Right: "The team taped all baseboards and patched throughout the unit."
        - The client does not need to know which person did which task. Speak for the crew as one unit.

        FORMAT:
        - Write in flowing paragraphs — no bullet points, no dashes, no numbered lists anywhere in the body.
        - Two paragraphs maximum: one for what was done today, one for what's planned tomorrow.
        - If there's a delay or issue, weave it naturally into the first paragraph.
        - Do not use section headers like "Work Completed" or "Planned for Tomorrow" — just write it.

        TONE:
        - Human and direct. Like a site manager giving a real update over email.
        - Use the actual language from the crew's check-in descriptions — their words, condensed into prose.
        - No padding, no filler. Every sentence should carry information.
        - No forced closings or pleasantries.

        BANNED phrases — never use:
        - "I hope this email finds you well"
        - "Please don't hesitate to reach out"
        - "Going forward" / "moving forward"
        - "It was a pleasure" / "At your earliest convenience"
        - "Thank you for your continued support" / "We appreciate your patience"
        - Any bullet points, dashes, or numbered lists in the body text

        Contact line at the bottom:
            Ahmad Ashrah | info@ashrahpainting.ca

        Respond with a JSON object containing three keys:
        {
          "subject":    "<email subject line>",
          "html_body":  "<full HTML email — use inline styles only, no external CSS>",
          "plain_body": "<plain text version of the same content>"
        }

        The HTML body must include:
        - An <h2> header: "Ashrah Painting — Daily Site Report"
        - Date and project address as a subheader
        - A colour-coded status badge inline <span>:
            green background (#2e7d32, white text) for "On Schedule"
            #f0ad4e background for "Minor Delay"
            #d9534f background for "Significant Delay"
        - The report body as <p> paragraphs only — no <ul>, no <li>, no <br>-separated lists
        - Crew on Site as a single line: e.g. "Crew on site: Weas Alshawakh, Ismael Al ali"
        - Footer with contact info in small grey text

        Respond only with valid JSON. Do not wrap in markdown fences.
    """).strip()

    def __init__(
        self,
        client: anthropic.Anthropic,
        smtp_host: str,
        smtp_port: int,
        user: str,
        password: str,
        from_email: str,
    ):
        self.client     = client
        self.smtp_host  = smtp_host
        self.smtp_port  = smtp_port
        self.user       = user
        self.password   = password
        self.from_email = from_email

    def _resolve_names(self, worker_ids: list[str], tracker: WorkforceTracker) -> list[str]:
        lookup = {w.worker_id: w.name for w in tracker.get_all()}
        return [lookup.get(wid, wid) for wid in worker_ids]

    def generate(self, report: DailyReport, tracker: WorkforceTracker) -> dict[str, str]:
        crew_names = self._resolve_names(report.crew_present, tracker)

        data = {
            "report_date":    report.report_date,
            "job_id":         report.job_id,
            "site_address":   report.site_address,
            "client_name":    report.client_name,
            "crew_present":   crew_names,
            "work_completed": report.work_completed,
            "work_planned":   report.work_planned,
            "issues":         report.issues or "None.",
            "photos_pending": report.photos_pending,
            "overall_status": report.overall_status,
        }

        with self.client.messages.stream(
            model=MODEL,
            max_tokens=4096,
            system=self.SYSTEM_PROMPT,
            messages=[{
                "role": "user",
                "content": f"Generate the daily report for:\n{json.dumps(data, indent=2)}"
            }],
        ) as stream:
            response = stream.get_final_message()

        raw = next((b.text for b in response.content if b.type == "text"), "{}")
        raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()

        try:
            return json.loads(raw)
        except json.JSONDecodeError as exc:
            print(f"[DailyReportSender] JSON parse error: {exc}")
            # Safe fallback — wrap raw output in minimal HTML
            fallback_subject = (
                f"Daily Site Report — {report.site_address} — {report.report_date}"
            )
            return {
                "subject":    fallback_subject,
                "html_body":  f"<pre style='font-family:monospace'>{raw}</pre>",
                "plain_body": raw,
            }

    def send(
        self,
        report_content: dict[str, str],
        to_email: str,
        cc_emails: list[str] | None = None,
    ) -> bool:
        subject    = report_content.get("subject",    "Daily Site Report")
        html_body  = report_content.get("html_body",  "")
        plain_body = report_content.get("plain_body", "")

        msg             = MIMEMultipart("alternative")
        msg["Subject"]  = subject
        msg["From"]     = self.from_email
        msg["To"]       = to_email
        if cc_emails:
            msg["Cc"] = ", ".join(cc_emails)

        msg.attach(MIMEText(plain_body, "plain"))
        msg.attach(MIMEText(html_body,  "html"))

        recipients = [to_email] + (cc_emails or [])

        try:
            with smtplib.SMTP(self.smtp_host, 587, timeout=20) as server:
                server.ehlo()
                server.starttls()
                server.login(self.user, self.password)
                server.sendmail(self.from_email, recipients, msg.as_string())
            print(f"[DailyReportSender] Sent to {to_email}")
            return True
        except Exception as exc:
            print(f"[DailyReportSender] SMTP error: {exc}")
            return False


# ---------------------------------------------------------------------------
# MODULE 7 — EMPLOYEE DAILY LOG SHEET
# ---------------------------------------------------------------------------

EXCEL_LOG_PATH = os.getenv("LUMIA_LOG_PATH", os.path.expanduser("~/Desktop/Lumia/lumia_employee_log.xlsx"))

HEADERS = [
    "Date", "Employee Name", "Job ID", "Site Address",
    "Tape & Covering", "Drop Sheets", "Patching Process",
    "Paint Execution", "Site Control", "Washing & Tool Care",
    "Avg Score", "Daily Summary", "Custom Scores", "Tomorrow's Plan", "Notes"
]

# Column widths
COL_WIDTHS = [14, 22, 14, 36, 16, 14, 16, 16, 14, 20, 12, 48, 36, 48, 30]

# Colours
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"   # white
ALT_ROW_BG  = "DCE6F1"   # light blue
SCORE_GOOD  = "C6EFCE"   # green  (8-10)
SCORE_MID   = "FFEB9C"   # yellow (5-7)
SCORE_LOW   = "FFC7CE"   # red    (1-4)

def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


class EmployeeLogSheet:
    """
    Creates or appends to an Excel workbook that logs each employee's
    daily entries: site worked, work done, self-score.
    One sheet per month (e.g. "Apr 2026"), with a summary sheet.
    """

    def __init__(self, path: str = EXCEL_LOG_PATH):
        self.path = path

    # ---- Internal helpers ---------------------------------------------------

    def _get_or_create_workbook(self) -> openpyxl.Workbook:
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        if os.path.exists(self.path):
            return openpyxl.load_workbook(self.path)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet
        return wb

    def _get_or_create_sheet(self, wb: openpyxl.Workbook, sheet_name: str):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        self._write_header_row(ws)
        return ws

    def _write_header_row(self, ws) -> None:
        header_font    = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
        header_fill    = PatternFill("solid", start_color=HEADER_BG)
        header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (heading, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=heading)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = _thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def _score_fill(self, score: int) -> PatternFill:
        if score >= 8:
            return PatternFill("solid", start_color=SCORE_GOOD)
        if score >= 5:
            return PatternFill("solid", start_color=SCORE_MID)
        return PatternFill("solid", start_color=SCORE_LOW)

    def _write_entry_row(self, ws, row: int, entry: EmployeeDailyEntry, alt: bool) -> None:
        cat_scores = [
            entry.tape_covering, entry.drop_sheets, entry.patching_process,
            entry.paint_execution, entry.site_control, entry.washing_tool_care,
        ]
        filled = [s for s in cat_scores if s > 0]
        avg    = round(sum(filled) / len(filled), 1) if filled else 0

        values = [
            entry.entry_date,
            entry.worker_name,
            entry.job_id,
            entry.site_address,
            entry.tape_covering or "—",
            entry.drop_sheets or "—",
            entry.patching_process or "—",
            entry.paint_execution or "—",
            entry.site_control or "—",
            entry.washing_tool_care or "—",
            avg if avg else "—",
            entry.work_description,
            entry.custom_scores or "—",
            entry.tomorrows_plan or "—",
            entry.notes,
        ]
        # Score columns are indices 4-10 (0-based) → col 5-11 (1-based)
        SCORE_COLS = {5, 6, 7, 8, 9, 10, 11}
        base_fill  = PatternFill("solid", start_color=ALT_ROW_BG) if alt else None

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.border    = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_idx in SCORE_COLS and isinstance(value, (int, float)):
                cell.fill      = self._score_fill(int(value))
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif base_fill:
                cell.fill = base_fill

        ws.row_dimensions[row].height = 50

    def _rebuild_summary(self, wb: openpyxl.Workbook) -> None:
        """Regenerate the Summary sheet from all monthly sheets."""
        SUMMARY = "Summary"
        if SUMMARY in wb.sheetnames:
            del wb[SUMMARY]

        ws = wb.create_sheet(SUMMARY, 0)

        # Title
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value     = "Lumia — Employee Log Summary"
        title_cell.font      = Font(name="Arial", bold=True, size=14, color=HEADER_FG)
        title_cell.fill      = PatternFill("solid", start_color=HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Sub-headers
        sub_headers = ["Month", "Entries", "Avg Self-Score", "Unique Employees", "Unique Sites", "", ""]
        sub_fill    = PatternFill("solid", start_color="2F5496")
        for col, h in enumerate(sub_headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill      = sub_fill
            c.alignment = Alignment(horizontal="center")
            c.border    = _thin_border()
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]

        row = 3
        for sheet_name in wb.sheetnames:
            if sheet_name == SUMMARY:
                continue
            src = wb[sheet_name]
            entries = src.max_row - 1   # subtract header row
            if entries <= 0:
                continue

            scores    = []
            employees = set()
            sites     = set()
            for data_row in src.iter_rows(min_row=2, values_only=True):
                if data_row[1]:
                    employees.add(data_row[1])
                if data_row[3]:
                    sites.add(data_row[3])
                if data_row[5] and isinstance(data_row[5], (int, float)):
                    scores.append(data_row[5])

            avg_score = f"=AVERAGE('{sheet_name}'!F2:F{src.max_row})" if scores else "—"

            row_data = [sheet_name, entries, avg_score, len(employees), len(sites), "", ""]
            alt = (row % 2 == 0)
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = Font(name="Arial", size=10)
                c.border    = _thin_border()
                c.alignment = Alignment(horizontal="center" if col != 1 else "left")
                if alt:
                    c.fill = PatternFill("solid", start_color=ALT_ROW_BG)
            row += 1

        ws.sheet_view.showGridLines = False

    # ---- Public API ---------------------------------------------------------

    def append_entries(self, entries: list[EmployeeDailyEntry]) -> str:
        """Append entries to the appropriate monthly sheet and save. Returns file path."""
        if not entries:
            return self.path

        wb = self._get_or_create_workbook()

        # Group entries by month sheet
        for entry in entries:
            try:
                dt         = datetime.strptime(entry.entry_date, "%Y-%m-%d")
                sheet_name = dt.strftime("%b %Y")    # e.g. "Apr 2026"
            except ValueError:
                sheet_name = "Misc"

            ws      = self._get_or_create_sheet(wb, sheet_name)
            next_row = ws.max_row + 1
            alt      = (next_row % 2 == 0)
            self._write_entry_row(ws, next_row, entry, alt)

        self._rebuild_summary(wb)
        wb.save(self.path)
        print(f"[EmployeeLogSheet] Saved {len(entries)} entry/entries → {self.path}")
        return self.path

    def get_today_entries_for_site(
        self, site_address: str, entry_date: str
    ) -> list[EmployeeDailyEntry]:
        """Return all logged entries for a given site and date (reads from Excel)."""
        if not os.path.exists(self.path):
            return []
        try:
            dt = datetime.strptime(entry_date, "%Y-%m-%d")
            sheet_name = dt.strftime("%b %Y")
        except ValueError:
            return []
        try:
            wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        except Exception:
            return []
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []

        site_lower = site_address.lower()
        results: list[EmployeeDailyEntry] = []
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            row_date = str(row[0]).strip()
            row_site = str(row[3] or "").strip().lower()
            if row_date != entry_date:
                continue
            # Match if either address contains the other (handles partial input)
            if site_lower not in row_site and row_site not in site_lower:
                continue
            def _int(v):
                return int(v) if isinstance(v, (int, float)) else 0
            results.append(EmployeeDailyEntry(
                entry_date=row_date,
                worker_id="",
                worker_name=str(row[1] or ""),
                job_id=str(row[2] or ""),
                site_address=str(row[3] or ""),
                tape_covering=_int(row[4]),
                drop_sheets=_int(row[5]),
                patching_process=_int(row[6]),
                paint_execution=_int(row[7]),
                site_control=_int(row[8]),
                washing_tool_care=_int(row[9]),
                self_score=_int(row[10]),
                work_description=str(row[11] or ""),
                notes=str(row[12] or ""),
            ))
        wb.close()
        return results


# ---------------------------------------------------------------------------
# MODULE 8 — SHARED SITE REPORT
# ---------------------------------------------------------------------------

OWNER_EMAIL = os.getenv("LUMIA_OWNER_EMAIL", "ahmad@ashrahpainting.ca")   # who gets the report

class SharedSiteReporter:
    """
    Finds sites where 2+ employees checked in on the same date,
    generates a combined Claude report for each, and emails it to the owner.
    """

    SYSTEM_PROMPT = textwrap.dedent("""
        You are writing an internal site summary report for Ahmad Ashrah,
        owner of Ashrah Painting.

        You are given the check-in entries of two or more employees who worked
        at the SAME site on the same day.

        Write a concise internal report covering:
        - Site address and date
        - Who was there
        - Combined summary of all work done
        - Average self-score and what it suggests about crew confidence
        - Any notes or flags worth Ahmad's attention

        Tone: direct and professional. Ahmad is the boss — no fluff.

        Respond with JSON only:
        {
          "subject":    "<email subject>",
          "html_body":  "<HTML email with inline styles>",
          "plain_body": "<plain text version>"
        }
    """).strip()

    def __init__(self, client: anthropic.Anthropic, smtp_host: str, smtp_port: int,
                 user: str, password: str, from_email: str):
        self.client     = client
        self.smtp_host  = smtp_host
        self.smtp_port  = smtp_port
        self.user       = user
        self.password   = password
        self.from_email = from_email

    def find_shared_sites(
        self, entries: list[EmployeeDailyEntry]
    ) -> dict[tuple[str, str], list[EmployeeDailyEntry]]:
        """Group entries by (date, site). Return only groups with 2+ employees."""
        groups: dict[tuple[str, str], list[EmployeeDailyEntry]] = {}
        for e in entries:
            key = (e.entry_date, e.site_address.strip().lower())
            groups.setdefault(key, []).append(e)
        return {k: v for k, v in groups.items() if len(v) >= 2}

    def _generate_report(self, entries: list[EmployeeDailyEntry]) -> dict[str, str]:
        payload = json.dumps([
            {
                "name":             e.worker_name,
                "site":             e.site_address,
                "date":             e.entry_date,
                "work_description": e.work_description,
                "self_score":       e.self_score,
                "notes":            e.notes,
            }
            for e in entries
        ], indent=2)

        try:
            with self.client.messages.stream(
                model=MODEL,
                max_tokens=2048,
                system=self.SYSTEM_PROMPT,
                messages=[{"role": "user", "content": payload}],
            ) as stream:
                response = stream.get_final_message()
        except Exception as exc:
            print(f"[SharedSiteReporter] API error: {exc}")
            return {}

        raw = next((b.text for b in response.content if b.type == "text"), "{}")
        raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {"subject": "Site Report", "html_body": raw, "plain_body": raw}

    def _send(self, content: dict[str, str], to_email: str) -> bool:
        msg            = MIMEMultipart("alternative")
        msg["Subject"] = content.get("subject", "Shared Site Report")
        msg["From"]    = self.from_email
        msg["To"]      = to_email
        msg.attach(MIMEText(content.get("plain_body", ""), "plain"))
        msg.attach(MIMEText(content.get("html_body",  ""), "html"))
        try:
            with smtplib.SMTP_SSL(self.smtp_host, self.smtp_port) as server:
                server.login(self.user, self.password)
                server.sendmail(self.from_email, [to_email], msg.as_string())
            return True
        except smtplib.SMTPException as exc:
            print(f"[SharedSiteReporter] SMTP error: {exc}")
            return False

    def run(self, entries: list[EmployeeDailyEntry], owner_email: str) -> int:
        """Generate and send a report for every shared site. Returns count sent."""
        groups = self.find_shared_sites(entries)
        if not groups:
            print("[SharedSiteReporter] No shared sites found.")
            return 0

        sent = 0
        for (report_date, site), group in groups.items():
            names = ", ".join(e.worker_name for e in group)
            print(f"[SharedSiteReporter] Shared site: {site} on {report_date} — {names}")
            content = self._generate_report(group)
            if content and self._send(content, owner_email):
                print(f"  → Report sent to {owner_email}: {content.get('subject','')}")
                sent += 1
        return sent


# ---------------------------------------------------------------------------
# ORCHESTRATOR — ties all 8 modules together
# ---------------------------------------------------------------------------

class Lumia:
    """
    Lumia — Ashrah Painting Operations Agent.
    Call run() to execute the full pipeline:
      1. Scan Zoho inbox for unread emails
      2. Extract structured job data from emails
      3. Print workforce availability snapshot
      4. Compare QC inspection reports (if provided)
      5. Assign available crews to open jobs
      6. Generate and email daily site reports to clients
      7. Record employee daily entries to Excel log
    """

    def __init__(self, demo: bool = False):
        if not ANTHROPIC_API_KEY:
            raise RuntimeError("ANTHROPIC_API_KEY environment variable not set.")
        if not demo and not ZOHO_PASSWORD:
            raise RuntimeError("ZOHO_PASSWORD environment variable not set. (Not required in --demo mode)")

        self.claude   = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        self.scanner  = EmailScanner(ZOHO_IMAP_HOST, ZOHO_IMAP_PORT, ZOHO_EMAIL, ZOHO_PASSWORD)
        self.analyzer = JobAnalyzer(self.claude)
        self.tracker  = WorkforceTracker()
        self.qc       = QCInspectorComparison(self.claude)
        self.assigner = SiteAssigner(self.claude, self.tracker)
        self.reporter = DailyReportSender(
            self.claude,
            ZOHO_SMTP_HOST,
            ZOHO_SMTP_PORT,
            ZOHO_EMAIL,
            ZOHO_PASSWORD,
            ZOHO_EMAIL,
        )
        self.log_sheet       = EmployeeLogSheet()
        self.emp_parser      = EmployeeEmailParser(self.claude)
        self.site_reporter   = SharedSiteReporter(
            self.claude, ZOHO_SMTP_HOST, ZOHO_SMTP_PORT,
            ZOHO_EMAIL, ZOHO_PASSWORD, ZOHO_EMAIL,
        )
        self._seed_workforce()

    def _seed_workforce(self) -> None:
        """Seed a realistic demo crew. Replace with a DB/file load in production."""
        workers = [
            Worker("W001", "Carlos Mendez",  ["exterior", "spray", "surface_prep"],        True,  certifications=["lead-safe"]),
            Worker("W002", "Jamal Brooks",   ["interior", "trim", "drywall"],               True),
            Worker("W003", "Sofia Reyes",    ["interior", "exterior", "spray"],             True,  certifications=["lead-safe"]),
            Worker("W004", "Derek Okafor",   ["exterior", "surface_prep"],                  True),
            Worker("W005", "Maria Castillo", ["interior", "trim"],                          True),
            Worker("W006", "Tony Nguyen",    ["exterior", "spray", "commercial"],           True,  certifications=["lead-safe", "lift_operator"]),
            Worker("W007", "Priya Sharma",   ["interior", "commercial", "drywall"],         True),
            Worker("W008", "James Porter",   ["exterior", "surface_prep", "caulking"],      True),
        ]
        for w in workers:
            self.tracker.add_worker(w)

    # ---- Pipeline steps -------------------------------------------------------

    def step1_scan_emails(self) -> list[EmailMessage]:
        print("\n[Step 1] Scanning Zoho inbox for unread messages...")
        messages = self.scanner.fetch_unread(max_count=5)
        print(f"         Found {len(messages)} unread message(s).")
        return messages

    def step2_analyze_jobs(self, messages: list[EmailMessage]) -> list[Job]:
        print("\n[Step 2] Analysing emails for job details...")
        if not messages:
            print("         No messages to analyse.")
            return []
        jobs = self.analyzer.analyze(messages)
        print(f"         Extracted {len(jobs)} job(s).")
        for j in jobs:
            print(f"           • {j.job_id}: {j.client_name} @ {j.site_address} [{j.status}]")
        return jobs

    def step2b_parse_employee_emails(self, messages: list[EmailMessage]) -> list[EmployeeDailyEntry]:
        print("\n[Step 2b] Parsing employee check-in emails...")
        if not messages:
            print("          No messages.")
            return []
        entries = self.emp_parser.parse_entries(messages)
        print(f"          Found {len(entries)} employee entry/entries.")
        return entries

    def step3_workforce_status(self) -> None:
        print("\n[Step 3] Current workforce status:")
        print(self.tracker.summary())

    def step4_qc_comparison(self, reports: list[QCReport]) -> dict[str, Any]:
        print("\n[Step 4] Running QC inspector comparison...")
        if len(reports) < 2:
            print("         Need ≥2 reports; skipping.")
            return {}
        result = self.qc.compare(reports)
        print(self.qc.format_comparison(result))
        return result

    def step5_assign_crews(self, jobs: list[Job]) -> list[SiteAssignment]:
        print("\n[Step 5] Assigning crews to open jobs...")
        assignments = self.assigner.assign(jobs)
        if assignments:
            print(self.assigner.summary())
        else:
            print("         No assignments made.")
        return assignments

    def step7_log_employee_entries(self, entries: list[EmployeeDailyEntry]) -> None:
        print("\n[Step 7] Writing employee daily log to Excel...")
        if not entries:
            print("         No entries to log.")
            return
        path = self.log_sheet.append_entries(entries)
        print(f"         Logged {len(entries)} entry/entries → {path}")

    def step6_send_daily_reports(self, daily_reports: list[DailyReport]) -> None:
        print("\n[Step 6] Generating and sending daily site reports...")
        for dr in daily_reports:
            print(f"         → {dr.client_name} / {dr.site_address}")
            content = self.reporter.generate(dr, self.tracker)
            sent    = self.reporter.send(content, to_email=dr.client_email)
            status  = "SENT" if sent else "FAILED"
            print(f"           [{status}] {content.get('subject', '')}")

    def step8_shared_site_reports(self, entries: list[EmployeeDailyEntry]) -> None:
        print("\n[Step 8] Checking for shared sites and sending report to owner...")
        count = self.site_reporter.run(entries, OWNER_EMAIL)
        if count == 0:
            print("         No shared-site reports sent.")

    # ---- Main entry point -----------------------------------------------------

    def run(
        self,
        demo_qc_reports: list[QCReport] | None = None,
        demo_daily_reports: list[DailyReport] | None = None,
        demo_employee_entries: list[EmployeeDailyEntry] | None = None,
        mark_emails_read: bool = False,
    ) -> None:
        print("=" * 60)
        print("  LUMIA — ASHRAH PAINTING OPERATIONS AGENT")
        print(f"  Run date : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 60)

        messages = self.step1_scan_emails()
        if mark_emails_read:
            for m in messages:
                self.scanner.mark_as_read(m.uid)

        jobs = self.step2_analyze_jobs(messages)
        email_entries = self.step2b_parse_employee_emails(messages)
        self.step3_workforce_status()

        if demo_qc_reports:
            self.step4_qc_comparison(demo_qc_reports)

        if jobs:
            self.step5_assign_crews(jobs)

        if demo_daily_reports:
            self.step6_send_daily_reports(demo_daily_reports)

        # Merge real email entries with any demo entries passed in
        all_entries = email_entries + (demo_employee_entries or [])

        self.step7_log_employee_entries(all_entries)
        self.step8_shared_site_reports(all_entries)

        print("\n[Lumia] Pipeline complete.\n")


# ---------------------------------------------------------------------------
# DEMO DATA — realistic fixtures for a dry run (no live email required)
# ---------------------------------------------------------------------------

def build_demo_data() -> tuple[list[QCReport], list[DailyReport], list[EmployeeDailyEntry]]:
    # Two inspectors reviewed the same kitchen repaint
    qc_reports = [
        QCReport(
            report_id="QC-2026-0401-A",
            job_id="JOB-2026-042",
            inspector_name="Rachel Kim",
            inspection_date="2026-04-02",
            items={
                "Kitchen ceiling":  "PASS — uniform coverage, no holidays",
                "Trim/baseboards":  "FAIL — brush marks visible near doorframe",
                "Accent wall":      "PASS",
                "Cabinet faces":    "PASS — two coats applied",
                "Surface prep":     "PASS — sanded and primed",
            },
            overall_rating="B+",
            raw_text="Minor touch-up needed at north doorframe trim. All other areas meet spec.",
        ),
        QCReport(
            report_id="QC-2026-0401-B",
            job_id="JOB-2026-042",
            inspector_name="Tom Bradley",
            inspection_date="2026-04-02",
            items={
                "Kitchen ceiling":  "PASS",
                "Trim/baseboards":  "PASS — acceptable finish",
                "Accent wall":      "FAIL — slight lap marks under raking light",
                "Cabinet faces":    "PASS",
                "Surface prep":     "PASS",
            },
            overall_rating="B",
            raw_text="Accent wall needs one more light coat to eliminate lap marks.",
        ),
    ]

    # One daily report ready to send
    daily_reports = [
        DailyReport(
            report_date=date.today().isoformat(),
            job_id="JOB-2026-039",
            site_address="1425 Lakeview Drive, Suite 200, Chicago IL 60614",
            client_name="Meridian Property Group",
            client_email=os.getenv("DEMO_CLIENT_EMAIL", "client@example.com"),
            crew_present=["W001", "W003", "W006"],
            work_completed=(
                "Completed surface prep and first coat on north and west exterior walls. "
                "Approximately 1,800 sq ft covered. Masking removed and scaffold repositioned "
                "to east elevation."
            ),
            work_planned=(
                "Second coat on north/west elevations. Begin surface prep on east elevation. "
                "Weather permitting, start east wall first coat."
            ),
            issues=(
                "Brief rain delay 10:30–11:15 AM. No material impact to schedule. "
                "Spray gun issue resolved on-site within 15 minutes."
            ),
            photos_pending=6,
            overall_status="On Schedule",
        ),
    ]

    today = date.today().isoformat()
    employee_entries = [
        EmployeeDailyEntry(
            entry_date=today, worker_id="W001", worker_name="Carlos Mendez",
            site_address="1425 Lakeview Drive, Suite 200, Chicago IL 60614",
            job_id="JOB-2026-039",
            work_description="Sprayed first coat on north and west exterior walls. Set up masking on east elevation.",
            self_score=8, notes="Ran out of masking tape around 2pm, borrowed from truck 2.",
        ),
        EmployeeDailyEntry(
            entry_date=today, worker_id="W003", worker_name="Sofia Reyes",
            site_address="1425 Lakeview Drive, Suite 200, Chicago IL 60614",
            job_id="JOB-2026-039",
            work_description="Back-rolled north wall after spray. Touched up window trim on west side.",
            self_score=9, notes="",
        ),
        EmployeeDailyEntry(
            entry_date=today, worker_id="W006", worker_name="Tony Nguyen",
            site_address="1425 Lakeview Drive, Suite 200, Chicago IL 60614",
            job_id="JOB-2026-039",
            work_description="Operated lift for upper sections of west wall. Repositioned scaffold after rain delay.",
            self_score=7, notes="Rain delay 10:30-11:15 AM slowed progress slightly.",
        ),
        EmployeeDailyEntry(
            entry_date=today, worker_id="W002", worker_name="Jamal Brooks",
            site_address="832 N. Michigan Ave, Unit 4B, Chicago IL 60611",
            job_id="JOB-2026-041",
            work_description="Cut in ceiling on master bedroom and hallway. Applied first coat to bedroom walls.",
            self_score=9, notes="Client requested slight colour adjustment — confirmed with supervisor.",
        ),
        EmployeeDailyEntry(
            entry_date=today, worker_id="W005", worker_name="Maria Castillo",
            site_address="832 N. Michigan Ave, Unit 4B, Chicago IL 60611",
            job_id="JOB-2026-041",
            work_description="Painted trim and baseboards in living room and dining room. Sanded and recoated one section.",
            self_score=8, notes="",
        ),
    ]

    return qc_reports, daily_reports, employee_entries


# ---------------------------------------------------------------------------
# REMINDER SENDER
# ---------------------------------------------------------------------------

def send_reminders() -> None:
    """
    Send a 4 PM daily check-in reminder to all employees listed in
    LUMIA_WORKER_EMAILS (comma-separated). Falls back to a printed warning
    if the env var is not set.
    """
    raw = os.getenv("LUMIA_WORKER_EMAILS", "")
    recipients = [e.strip() for e in raw.split(",") if e.strip()]

    if not recipients:
        print("[Reminder] No employee emails found. Set LUMIA_WORKER_EMAILS=email1,email2,...")
        return

    if not ZOHO_PASSWORD:
        print("[Reminder] ZOHO_PASSWORD not set — cannot send reminders.")
        return

    subject  = "Lumia — Daily Check-In Reminder"
    plain    = (
        "Hi team,\n\n"
        "This is your 4:00 PM reminder to send in your daily check-in.\n\n"
        "Send an email to ahmad@ashrahpainting.ca with:\n"
        "  Subject: Daily Check-In\n\n"
        "  1. Your name\n"
        "  2. The site address you worked at today\n"
        "  3. What you did\n"
        "  4. Your self-score out of 10\n\n"
        "Example:\n"
        "  \"Hi, this is Carlos. I worked at 123 Main St today.\n"
        "   Painted the exterior front wall, two coats. I'd give myself an 8/10.\"\n\n"
        "Thanks,\nLumia — Ashrah Painting"
    )
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto">
      <h2 style="background:#1F3864;color:#fff;padding:16px;margin:0;border-radius:6px 6px 0 0">
        Lumia — Daily Check-In Reminder
      </h2>
      <div style="border:1px solid #ddd;border-top:none;padding:20px;border-radius:0 0 6px 6px">
        <p>Hi team,</p>
        <p>This is your <strong>4:00 PM reminder</strong> to send in your daily check-in.</p>
        <p>Send an email to
           <a href="mailto:ahmad@ashrahpainting.ca" style="color:#1F3864;font-weight:bold">ahmad@ashrahpainting.ca</a>
           with subject <strong>"Daily Check-In"</strong> and include:</p>
        <ol>
          <li>Your <strong>name</strong></li>
          <li>The <strong>site address</strong> you worked at today</li>
          <li><strong>What you did</strong></li>
          <li>Your <strong>self-score</strong> out of 10</li>
        </ol>
        <div style="background:#f5f5f5;border-left:4px solid #1F3864;padding:12px;margin:16px 0;font-style:italic">
          "Hi, this is Carlos. I worked at 123 Main St today.
           Painted the exterior front wall, two coats. I'd give myself an 8/10."
        </div>
        <p style="color:#888;font-size:12px;margin-top:24px">
          Lumia — Ashrah Painting Operations Agent
        </p>
      </div>
    </div>
    """

    sent = 0
    try:
        with smtplib.SMTP_SSL(ZOHO_SMTP_HOST, ZOHO_SMTP_PORT) as server:
            server.login(ZOHO_EMAIL, ZOHO_PASSWORD)
            for to_email in recipients:
                msg            = MIMEMultipart("alternative")
                msg["Subject"] = subject
                msg["From"]    = ZOHO_EMAIL
                msg["To"]      = to_email
                msg.attach(MIMEText(plain, "plain"))
                msg.attach(MIMEText(html,  "html"))
                server.sendmail(ZOHO_EMAIL, [to_email], msg.as_string())
                print(f"[Reminder] Sent to {to_email}")
                sent += 1
    except smtplib.SMTPException as exc:
        print(f"[Reminder] SMTP error: {exc}")

    print(f"[Reminder] Done — {sent}/{len(recipients)} reminder(s) sent.")


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    if "--remind" in sys.argv:
        send_reminders()
        sys.exit(0)

    demo_mode = "--demo" in sys.argv

    agent = Lumia(demo=demo_mode)

    if demo_mode:
        print("\n[Demo mode] Using pre-built fixtures — no live Zoho connection.\n")
        qc_rpts, daily_rpts, emp_entries = build_demo_data()
        agent.run(
            demo_qc_reports=qc_rpts,
            demo_daily_reports=daily_rpts,
            demo_employee_entries=emp_entries,
            mark_emails_read=False,
        )
    else:
        # Full live run — reads real inbox, sends real emails
        agent.run(mark_emails_read=True)
